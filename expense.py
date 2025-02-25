import streamlit as st
import pandas as pd
import os
from openpyxl import Workbook, load_workbook
import matplotlib.pyplot as plt
import seaborn as sns

# Constants
FILE_NAME = "expense.xlsx"
REQUIRED_COLUMNS = ["Category", "Amount", "Date", "Description"]

# Function to initialize Excel file if it doesn't exist
def initialize_excel_file():
    if not os.path.exists(FILE_NAME):
        wb = Workbook()
        ws = wb.active
        ws.append(REQUIRED_COLUMNS)
        wb.save(FILE_NAME)

# Function to load data from the Excel file and clean empty rows
def load_data_from_excel():
    initialize_excel_file()
    data = pd.read_excel(FILE_NAME)
    
    # Remove rows where any of the required columns are NaN
    data = data.dropna(subset=REQUIRED_COLUMNS)
    
    # Reset the index after dropping rows
    data = data.reset_index(drop=True)
    
    return data

# Function to add data to the Excel file
def add_data_to_excel(category, amount, date, description):
    initialize_excel_file()
    # Use a temporary file to handle potential locks
    temp_file = "temp_expense.xlsx"
    
    try:
        # Load the workbook
        wb = load_workbook(FILE_NAME)
        ws = wb.active
        
        # Append new data
        ws.append([category, amount, date, description])
        
        # Save to a temporary file and replace the original file
        wb.save(temp_file)
        os.replace(temp_file, FILE_NAME)
    except PermissionError:
        st.error("The file 'expense.xlsx' is currently open. Please close it and try again.")
    except Exception as e:
        st.error(f"An error occurred: {e}")
    finally:
        # Clean up temporary file if it exists
        if os.path.exists(temp_file):
            os.remove(temp_file)

# Initialize Excel file
initialize_excel_file()

# Page Title
st.title("Personal Expense Tracker")

# Sidebar Section
with st.sidebar:
    st.header("Add Section")
    
    # Category Dropdown
    category_options = ["Food", "Transport", "Entertainment", "Utilities", "Healthcare", "Other"]
    category = st.selectbox("Select Category", category_options)
    
    # Input fields for manual entry
    date = st.date_input("Date")
    amount = st.number_input("Amount", min_value=0.0, step=0.01)
    description = st.text_area("Description")
    
    # Add Expense Button
    if st.button("Add Expense"):
        if category and amount > 0 and description:
            add_data_to_excel(category, amount, date, description)
            st.success("Expense added successfully!")
        else:
            st.error("Please fill in all fields before adding an expense.")
    
    # OR Separator
    st.markdown("### OR ###")
    
    # File Upload Section
    st.subheader("Upload Expenses File")
    uploaded_file = st.file_uploader("Upload your file (CSV or Excel)", type=["csv", "xlsx"])
    
    if uploaded_file:
        try:
            # Read the uploaded file
            if uploaded_file.name.endswith(".csv"):
                uploaded_data = pd.read_csv(uploaded_file)
            else:
                uploaded_data = pd.read_excel(uploaded_file)
            
            # Check for required columns
            if all(col in uploaded_data.columns for col in REQUIRED_COLUMNS):
                # Clean the uploaded data by dropping empty rows
                uploaded_data = uploaded_data.dropna(subset=REQUIRED_COLUMNS)
                uploaded_data = uploaded_data.reset_index(drop=True)
                
                st.session_state["uploaded_data"] = uploaded_data
                st.success("File uploaded successfully!")
            else:
                st.error(f"The uploaded file must contain the following columns: {REQUIRED_COLUMNS}")
        except Exception as e:
            st.error(f"Error reading the file: {e}")

# Main Section
st.header("Expenses Table")

# Load data from file or session state
if "uploaded_data" in st.session_state:
    # Display uploaded file data
    data = st.session_state["uploaded_data"]
    st.info("Showing data from the uploaded file.")
else:
    # Load data from the Excel file
    data = load_data_from_excel()
    st.info("Showing data from the saved file.")

# Display the expenses table
st.dataframe(data, use_container_width=True)

# Visualization Section
st.header("Visualize Expenses")
if st.button("Visualize"):
    if not data.empty:
        # Expense by Category
        st.subheader("Expense by Category")
        category_expense = data.groupby("Category")["Amount"].sum().reset_index()
        fig1, ax1 = plt.subplots()
        ax1.pie(category_expense["Amount"], labels=category_expense["Category"], autopct="%1.1f%%", startangle=90)
        ax1.axis("equal")
        st.pyplot(fig1)
        
        # Monthly Expense Trend
        st.subheader("Monthly Expense Trend")
        data["Date"] = pd.to_datetime(data["Date"])
        data["Month"] = data["Date"].dt.to_period("M").astype(str)
        monthly_expense = data.groupby("Month")["Amount"].sum().reset_index()
        fig2, ax2 = plt.subplots()
        sns.lineplot(data=monthly_expense, x="Month", y="Amount", marker="o", ax=ax2)
        ax2.set_title("Monthly Expense Trend")
        plt.xticks(rotation=45)
        st.pyplot(fig2)
        
        # Bar Chart of Categories
        st.subheader("Bar Chart of Expenses by Category")
        fig3, ax3 = plt.subplots()
        sns.barplot(data=category_expense, x="Category", y="Amount", palette="viridis", ax=ax3)
        ax3.set_title("Expenses by Category")
        plt.xticks(rotation=45)
        st.pyplot(fig3)
    else:
        st.warning("No data available to visualize. Please add expenses first.")
